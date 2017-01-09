from openpyxl import Workbook
from openpyxl.styles import Font
#MR Equation (100/(100+MR)) or 2 - (100/(100+MR))
#  MR = MR * .65 if void staff is present
#Damage = ((Base dmg + Ratio)[*Missing HP mult]*MR Multiplier
#Current HP = Total HP - Damage
#Q - 80/130/180/230/280 (50% AP) [20/22/24/26/28% Shred]
#E - 60/110/160/210/260 (70% AP)
#R - 100/140/180 (25% AP) | 200/280/360 (50% AP) <= Execute (below 40% HP)
#Generate results for 30, 40, 50, 75, 100, 150, 200, MR
#Ashe HP 528-1871 (1221 at 11) 30 MR
#Alistar HP 613-2415 (1544 at 11) +1.25 Mr per level (43 at 11)
#Masteries:-
#%7.5 Consistent, %2.5 with oppression
#~6 extra pen at lvl 11, ~8 at lvl 16
#Liandrys - 2% current HP, 4% with slow

#Red
font = Font(color="FF0000")

def PrepSheet(Health, AP, Mpen, Base, MRignore, shots, Liandrys, Rylais, wb):
    ws = wb.create_sheet("Sheet")
    #Set up 1st row
    tmp = "{0}={1}"
    ws['A1'] = tmp.format("Health",Health)
    ws['B1'] = tmp.format("AP",AP)
    ws['C1'] = tmp.format("Mpen",Mpen)
    ws['D1'] = tmp.format("Base",Base)
    ws['E1'] = tmp.format("MRignore",MRignore)
    ws['F1'] = tmp.format("Liandrys",Liandrys)
    ws['G1'] = tmp.format("Rylais",Rylais)
    #Setup 2nd row (shots)
    ws['A2'] = "MR-v Shot#->"
    for i in range(1, shots+1):
        ws.cell(row=2, column=i+1, value=i)
    Row = 3
    for i in range(30,201,10):     
        if (i > 50 and i < 70) or (i > 70 and i < 100) or (i > 100 and i < 150) or (i > 150 and i < 200):
            continue #Only keep the important ones
        MR = i
        MR = MR * MRignore #Change only if voidstaff or Q is present
        MR = MR - Mpen
        MRActual = MR
        tmp_str = str(i) + " (" + str(MRActual) + ")"
        ws.cell(row=Row,column=1,value=tmp_str)
        Row = Row + 1

    return ws
    

def Barrage (Health, AP, Mpen, Base, MRignore, shots, Liandrys, Rylais, wb):
    ws = PrepSheet(Health, AP, Mpen, Base, MRignore, shots, Liandrys, Rylais, wb)
    Column = 2
    if Base == 1:
        Base = 100
    elif Base == 2:
        Base = 140
    elif Base == 3:
        Base = 180

    HP = []
    for i in range(0,shots+2):
        HP.append(Health)
    MRmult = 1 #How much damage we are actually doing to target
    Hpcount = 0 #is this 30, 40, 50 etc
    MRActual = 0

    for shot in range(1,shots+1):
        Missinghp = Health - HP[Hpcount]
        Missingmult = Missinghp/Health
        Missingmult = .83 * Missingmult
        tmp = "Shot {0}:"
        print(tmp.format(shot))
        Row = 3
        #For each stage of MR
        for i in range(30,201,10):
            if (i > 50 and i < 70) or (i > 70 and i < 100) or (i > 100 and i < 150) or (i > 150 and i < 200):
                continue #Only keep the important ones
            #Lets get into the actual calculations!
            ExecuteFlag = 0
            Current = HP[Hpcount]
            MR = i
            MR = MR * MRignore #Change only if voidstaff or Q is present
            MR = MR - Mpen
            MRActual = MR
            if MR >= 0:
                MRmult = 100.0/(100+MR)
            else: #Negative MR
                MRmult = 2-(100.0/(100-MR))

            if (Current/Health > .4):
                Damage = (Base + (AP*.25))*(1+Missingmult)*MRmult
            else: #Execute
                Damage = (Base*2 + (AP*.5))*MRmult
                ExecuteFlag = 1

            Damage = Damage * 1.075 #Regular Mastery Amp (assuming assassin)
            if shot > 1 and Rylais == 1:
                Damage = Damage * 1.025 #Opressor only works after 1st shot?
            Damage = round(Damage)

            HP[Hpcount] = HP[Hpcount] - Damage
            tmp = "MR:{0}({3}) | Damage:{1} | Remaining: {2} {4} [Dealt:{5}]"
            if HP[Hpcount] <= 0:
                filler = "dead"
            else:
                filler = HP[Hpcount]
            exfiller = ""
            if ExecuteFlag == 1:
                exfiller = "!"
            else:
                exfiller = ""
            print (tmp.format(i,Damage,filler, MRActual, exfiller, Health-HP[Hpcount]))

            Cell = str(Damage) + "(" + str(filler) + ")"
            #Check if we need to calculate Liandry
            if Liandrys == 1: #TODO - Does not calculate to total hp?
                #4% current hp, 1.6 seconds = 6.4% current HP burn  
                Burn = HP[Hpcount] * 0.064
                Burn = Burn * MRmult
                Burn = round(Burn)
                temp_string = "Liandry's burned for:{0} [{1} Remaining]"
                if HP[Hpcount] > 0:
                    print(temp_string.format(Burn,HP[Hpcount]-Burn))
                HP[Hpcount] = HP[Hpcount] - Burn
                Cell = Cell + "\n[" + str(Burn) + "]"
            c = ws.cell(row=Row ,column= Column,value= Cell)
            if ExecuteFlag == 1:
                c.font = font
            Hpcount = Hpcount + 1
            Row = Row + 1      
        Hpcount = 0
        Column = Column + 1

